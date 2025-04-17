import argparse
import numpy as np
import torch
from torch import optim
import pandas as pd
from utils_z import MetaAcc, pretty_print_ly
from utils import CosineLR

from choose_dataset import init_dataset
from algorithms import algorithm_builder

import os
from tqdm import tqdm
from openpyxl import load_workbook, Workbook

algorithmMap = {
    "irmv1": "IRMV1",
    "infer_irmv1": "ZIN",
    "irmv1_tvl1": "IRM-TV-L1",
    "infer_irmv1_tvl1": "ZIN-TV-L1",
    "irmv1_multi_class": "IRMV1",
    "infer_irmv1_multi_class": "ZIN",
    "irmv1_multi_class_tvl1": "IRM-TV-L1",
    "infer_irmv1_multi_class_tvl1": "ZIN-TV-L1",
}


def writeResult(flags, result: list):

    column_names = [
        "step",
        "iter",
        "lr",
        "lr2",
        "penalty_weight_anneal",
        "Final train acc",
        "Final test acc",
        "Worst test acc",
        "Current step",
    ]
    filename = f"{flags.dataset}_{flags.irm_type}.xlsx"
    filePath = f"result/{filename}"

    try:

        wb = load_workbook(filename=filePath)
        ws = wb.active
    except FileNotFoundError:

        wb = Workbook()
        ws = wb.active

        for col, col_name in enumerate(column_names, 1):
            ws.cell(row=1, column=col).value = col_name

    start_row = ws.max_row + 1

    row = [
        flags.steps,
        flags.penalty_anneal_iters,
        flags.lr,
        flags.lr2,
        flags.penalty_weight_anneal,
    ] + result

    for idx, value in enumerate(row):
        ws.cell(row=start_row, column=idx + 1, value=value)

    wb.save(filename=filePath)


parser = argparse.ArgumentParser(description="ZIN")
parser.add_argument("--aux_num", type=int, default=7)
parser.add_argument("--batch_size", type=int, default=1024)
parser.add_argument("--seed", type=int, default=0)
parser.add_argument("--classes_num", type=int, default=2)
parser.add_argument(
    "--dataset",
    type=str,
    default="mnist",
    choices=[
        "celebaz_feature",  # 1
        "logit",
        "logit_z",  # 1
        "logit_2z",
        "house_price",  # 1
        "landcover",  # 1
        "adult",
        "mnist",
    ],
)
parser.add_argument("--opt", type=str, default="adam", choices=["adam", "sgd"])
parser.add_argument("--l2_regularizer_weight", type=float, default=0.001)
parser.add_argument("--print_every", type=int, default=1)
parser.add_argument("--dim_inv", type=int, default=2)
parser.add_argument("--dim_spu", type=int, default=10)
parser.add_argument("--data_num_train", type=int, default=2000)
parser.add_argument("--data_num_test", type=int, default=2000)
parser.add_argument("--lr", type=float, default=0.001)
parser.add_argument("--lr2", type=float, default=0.001)  
parser.add_argument(
    "--env_type", default="linear", type=str, choices=["2_group", "cos", "linear"]
)
parser.add_argument(
    "--irm_type",
    default="irmv1",
    type=str,
    choices=[
        "erm",
        "lff",
        "eiil",
        "irmv1",
        "irmv1_tvl1",
        "infer_irmv1",
        "infer_irmv1_tvl1",
        "infer_irmv1_multi_class",
        "infer_irmv1_multi_class_tvl1",
        "irmv1_multi_class",
        "irmv1_multi_class_tvl1",
    ],
)
parser.add_argument("--n_restarts", type=int, default=10)
parser.add_argument("--image_scale", type=int, default=64)
parser.add_argument("--hidden_dim", type=int, default=16)
parser.add_argument("--hidden_dim_infer", type=int, default=16)
parser.add_argument("--cons_train", type=str, default="0.999_0.7")
parser.add_argument("--cons_test", type=str, default="0.999_0.001")
parser.add_argument("--num_classes", type=int, default=2)
parser.add_argument("--z_class_num", type=int, default=4)
parser.add_argument("--noise_ratio", type=float, default=0.1)
parser.add_argument("--penalty_anneal_iters", type=int, default=200)
parser.add_argument("--penalty_weight_anneal", type=int, default=200)  
parser.add_argument("--penalty_weight", type=float, default=1.0)
parser.add_argument("--steps", type=int, default=501)
parser.add_argument("--scheduler", type=int, default=0)
# adult
parser.add_argument("--envs_num_train", type=int)
parser.add_argument("--envs_num_test", type=int)

flags = parser.parse_args()


saveSpan = int(flags.steps * 0.01) + 1
saveBar = 100
if flags.dataset == "mnist":
    saveSpan = 1
    saveBar = 0


RESULT_PATH = "result"
os.makedirs(RESULT_PATH, exist_ok=True)


print("batch_size is", flags.batch_size)
torch.autograd.set_detect_anomaly(True)
torch.manual_seed(flags.seed)
if flags.dataset == "landcover":
    np.random.seed(flags.seed + 111)  # to be consistent with in-N-out.
if not flags.dataset == "adult":
    flags.cons_ratio = "_".join([flags.cons_train, flags.cons_test])
    flags.envs_num_train = len(flags.cons_train.split("_"))
    flags.envs_num_test = len(flags.cons_test.split("_"))
    assert flags.envs_num_test + flags.envs_num_train == len(
        flags.cons_ratio.split("_")
    )

for k, v in sorted(vars(flags).items()):
    print("\t{}: {}".format(k, v))

final_train_accs = []
final_test_accs = []

## add by w wang
final_test_accs_worst = []
lr = flags.lr
lr2 = flags.lr2
for restart in range(1):
    print("Restart", restart)

    (
        dp,
        mlp,
        mlp2,
        test_batch_num,
        train_batch_num,
        val_batch_num,
        test_batch_fetcher,
        mean_nll,
        mean_accuracy,
        eval_acc,
    ) = init_dataset(flags)
    if mean_accuracy.__name__ == "mean_accuracy_reg":
        best_result_record = {
            "final": 10000,
            "worst": 10000,
            "final_epoch": -1,
            "worst_epoch": -1,
        }
    else:
        best_result_record = {
            "final": 0.0,
            "worst": 0,
            "final_epoch": -1,
            "worst_epoch": -1,
        }
    # if flags.opt == "adam":
    optimizer = optim.Adam(mlp.parameters(), lr=flags.lr)
    #     optimizer2 = optim.Adam(mlp2.parameters(), lr=flags.lr2)
    # elif flags.opt == "sgd":
    # optimizer = optim.SGD(mlp.parameters(), momentum=0.9, lr=flags.lr)
    # optimizer2 = optim.SGD(mlp2.parameters(), momentum=0.9, lr=flags.lr2)

    scale = torch.tensor(1.0).cuda().requires_grad_()
    algo = algorithm_builder(flags, dp)

    if flags.dataset == "house_price":
        meta_acc_test = MetaAcc(
            env=dp.envs_num_test, acc_measure=mean_accuracy, acc_type="test"
        )
    elif flags.dataset == "landcover":
        meta_acc_test = MetaAcc(env=2, acc_measure=mean_accuracy, acc_type="test")
    else:
        meta_acc_test = MetaAcc(env=0, acc_measure=mean_accuracy, acc_type="test")

    pretty_print_ly(
        ["step", "train loss", "train penalty", "penalty weight", "train acc"]
        + meta_acc_test.acc_fields
    )
    best_results = {"val_acc": 0.0, "epoch": -1, "train_acc": 0.0, "test_acc": 0.0}
    redRecord = []
    blueRecord = []
    fidelity = []
    residual = []
    residual2 = []
    oldData = []
    oldData2 = []
    psi = []
    mlp2Grad = []
    for step in tqdm(range(flags.steps)):
        mlp.train()
        mlp2.train()

        # lr2*=0.1/(step+1)
        for batch in range(train_batch_num):
            batch_data = dp.fetch_train()
            train_x, train_y, train_z, train_g, train_c, train_invnoise = batch_data

            # calculate train loss for different algorithms and datasets

            train_nll, train_penalty, penalty_weight = algo(
                batch_data, step, mlp, mlp2, scale, mean_nll=mean_nll
            )

            weight_norm = torch.tensor(0.0).cuda()
            for w in mlp.parameters():
                weight_norm += w.norm().pow(2)

            loss = train_nll.clone()
            fidelity_temp = flags.l2_regularizer_weight * weight_norm
            loss += fidelity_temp
            fidelity.append(fidelity_temp.item())
            penalty_weight_flag = (
                # 1.0 if step >= flags.penalty_anneal_iters else 0.0
                flags.penalty_weight_anneal
                if step >= flags.penalty_anneal_iters
                else 0.0
            )
            if flags.irm_type == "erm":
                penalty_weight = 0
            train_penalty = torch.max(torch.tensor(0.0).cuda(), train_penalty.cuda())
            red = penalty_weight_flag * flags.penalty_weight * train_penalty
            loss += red
            redRecord.append(red.item())
            blueRecord.append(loss.item())
            # loss=loss/(1+penalty_weight)
            # if penalty_weight > 1.0:
            #     loss /= 1.0 + penalty_weight

            optimizer.zero_grad()
            # optimizer2.zero_grad()
            loss.backward(retain_graph=True)
            optimizer.step()
            # for param_mlp in mlp.parameters():
            #     if torch.mean(param_mlp.grad) != 0:
            #         t = flags.lr * param_mlp.grad / ((step + 1)**1.0002) / (param_mlp.grad.norm())
            #         param_mlp.data = param_mlp.data + t
            #     else:
            #         break
            for param_mlp2 in mlp2.parameters():
                if param_mlp2.grad != None and torch.mean(param_mlp2.grad) != 0:
                    t = (
                        lr2
                        * param_mlp2.grad
                        / ((step + 1) ** 1.0001)
                        / (param_mlp2.grad.norm())
                    )
                    param_mlp2.data = param_mlp2.data + t
                else:
                    break
            # for param_mlp2 in mlp2.parameters():
            #     if param_mlp2.grad!=None:
            #         param_mlp2.data = param_mlp2.data + flags.lr2 * param_mlp2.grad
            # phi
            newData = []
            for param_mlp in mlp.parameters():
                newData.append(param_mlp.data.detach().cpu())
            if len(oldData) == 0:
                oldData = newData
            temp = [abs(oldData[i] - newData[i]) for i in range(len(newData))]
            temp = [np.linalg.norm(item) for item in temp]
            residualTemp = np.linalg.norm(temp)
            residual.append(residualTemp)
            oldData = newData
            # psi
            newData2 = []
            newGrad = []

            for param_mlp2 in mlp2.parameters():
                if param_mlp2.grad != None:
                    newData2.append(param_mlp2.data.detach().cpu())
                    newGrad.append(np.linalg.norm(param_mlp2.grad.detach().cpu()))
            if len(oldData2) == 0:
                oldData2 = newData2
            mlp2Grad.append(np.linalg.norm(newGrad))
            psi.append(np.linalg.norm([np.linalg.norm(data) for data in newData2]))
            temp = [abs(oldData2[i] - newData2[i]) for i in range(len(newData2))]
            temp = [np.linalg.norm(item) for item in temp]
            residualTemp = np.linalg.norm(temp)
            residual2.append(residualTemp)
            oldData2 = newData2

            # scheduler2.step()

        if step % flags.print_every == 0:
            mlp.eval()

            # train_accs = []
            # for ii in range(train_batch_num):
            #     train_x, train_y, train_z, train_g, train_c, train_invnoise = (
            #         dp.fetch_train()
            #     )
            #     train_logits = mlp(train_x)
            #     train_acc, train_minacc, train_majacc = eval_acc(
            #         train_logits, train_g, train_c
            #     )
            #     train_accs.append(train_acc)
            # train_acc = sum(train_accs) / len(train_accs)

            meta_acc_test.clear()
            for ii in range(test_batch_num):
                test_x, test_y, test_z, test_g, test_c, test_invnoise = (
                    test_batch_fetcher()
                )
                test_logits = mlp(test_x)
                meta_acc_test.process_batch(test_g, test_logits)
            meta_acc_test_res = meta_acc_test.meta_acc

            # if val_batch_num:
            #     val_accs = []
            #     for ii in range(val_batch_num):
            #         val_x, val_y, val_z, val_g, val_c, val_invnoise = dp.fetch_val()
            #         val_logits = mlp(val_x)
            #         val_acc, val_minacc, val_majacc = eval_acc(val_logits, val_g, val_c)
            #         val_accs.append(val_acc)
            #     val_acc = sum(val_accs) / len(val_accs)
            #     if val_acc > best_results["val_acc"]:
            #         best_results["val_acc"] = val_acc
            #         best_results["epoch"] = step
            #         best_results["train_acc"] = train_acc
            #         best_results["test_acc"] = meta_acc_test_res["test_acc"]
        result = [
            0,
            np.mean(meta_acc_test_res["test_acc"].detach().cpu().numpy()),
            np.mean(meta_acc_test_res["test_acc_worst"].detach().cpu().numpy()),
            step,
        ]
        if step % saveSpan == 0 and step >= saveBar:
            writeResult(flags, result)
        if mean_accuracy.__name__ == "mean_accuracy_reg":
            if result[1] < best_result_record["final"]:
                best_result_record["final"] = result[1]
                best_result_record["final_epoch"] = step
            if result[2] < best_result_record["worst"]:
                best_result_record["worst"] = result[2]
                best_result_record["worst_epoch"] = step
        else:
            if result[1] > best_result_record["final"]:
                best_result_record["final"] = result[1]
                best_result_record["final_epoch"] = step
            if result[2] > best_result_record["worst"]:
                best_result_record["worst"] = result[2]
                best_result_record["worst_epoch"] = step
    t = flags.dataset.split("_")
    dataset = " ".join(t)
    filePath = f"figure_data/{dataset}_{algorithmMap[flags.irm_type]}.csv"
    df = pd.DataFrame(
        {
            "redRecord": redRecord,
            "blueRecord": blueRecord,
            "phi_residual": residual,
            "psi_residual": residual2,
            "fidelity": fidelity,
            "mlp2Grad": mlp2Grad,
            "psi": psi,
        }
    )
    df.to_csv(filePath, index=False)
    BEST_PATH = "best_result"
    bestResultPath = f"{BEST_PATH}/{flags.dataset}_{algorithmMap[flags.irm_type]}.csv"
    if not os.path.exists(bestResultPath):
        df = pd.DataFrame(
            {
                "final": [best_result_record["final"]],
                "worst": [best_result_record["worst"]],
                "final_epoch": [best_result_record["final_epoch"]],
                "worst_epoch": [best_result_record["worst_epoch"]],
            }
        )
        df.to_csv(bestResultPath, index=False)
    else:
        df = pd.read_csv(bestResultPath)
        new = pd.DataFrame(
            {
                "final": [best_result_record["final"]],
                "worst": [best_result_record["worst"]],
                "final_epoch": [best_result_record["final_epoch"]],
                "worst_epoch": [best_result_record["worst_epoch"]],
            }
        )
        df = pd.concat(
            [df, new],
            ignore_index=True,
        )
        df.to_csv(bestResultPath, index=False)
    # if val_batch_num:
    #     print("IID Test:")
    #     print(val_acc)
    #     print("best results:")
    #     print(best_results)
