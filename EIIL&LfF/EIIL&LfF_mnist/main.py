import argparse
import numpy as np
import torch
from torch import optim
import os

from utils_z import MetaAcc, pretty_print_ly
from utils import CosineLR
import pandas as pd
from choose_dataset import init_dataset
from algorithms import algorithm_builder
from tqdm import tqdm
from openpyxl import load_workbook, Workbook

algorithmMap = {
    "irmv1": "IRMV1",
    "infer_irmv1": "ZIN",
    "irmv1_tvl1": "IRM-TV-L1",
    "infer_irmv1_tvl1": "ZIN-TV-L1",
    "irmv1_multi_class": "IRMV1",
    "infer_irmv1_multi_class": "ZIN",
    "irmv1_multi_class_tvl1": "IRM-TV-L1",
    "infer_irmv1_multi_class_tvl1": "ZIN-TV-L1",
    "eiil": "EIIL",
    "lff": "LFF",
}

parser = argparse.ArgumentParser(description="ZIN")
parser.add_argument("--aux_num", type=int, default=7)
parser.add_argument("--batch_size", type=int, default=1024)
parser.add_argument("--seed", type=int, default=0)
parser.add_argument("--classes_num", type=int, default=2)
parser.add_argument(
    "--dataset",
    type=str,
    default="mnist",
    choices=[
        "celebaz_feature",
        "logit",
        "logit_z",
        "logit_2z",
        "house_price",
        "landcover",
        "mnist",
    ],
)
parser.add_argument("--opt", type=str, default="adam", choices=["adam", "sgd"])
parser.add_argument("--l2_regularizer_weight", type=float, default=0.001)
parser.add_argument("--print_every", type=int, default=1)
parser.add_argument("--dim_inv", type=int, default=2)
parser.add_argument("--dim_spu", type=int, default=10)
parser.add_argument("--data_num_train", type=int, default=2000)
parser.add_argument("--data_num_test", type=int, default=2000)
parser.add_argument("--lr", type=float, default=0.001)
parser.add_argument(
    "--env_type", default="linear", type=str, choices=["2_group", "cos", "linear"]
)
parser.add_argument(
    "--irm_type",
    default="irmv1",
    type=str,
    choices=[
        "erm",
        "lff",
        "eiil",
        "irmv1",
        "irmv1_tvl1",
        "infer_irmv1",
        "infer_irmv1_tvl1",
        "infer_irmv1_multi_class",
        "infer_irmv1_multi_class_tvl1",
        "irmv1_multi_class",
        "irmv1_multi_class_tvl1",
    ],
)
parser.add_argument("--n_restarts", type=int, default=10)
parser.add_argument("--image_scale", type=int, default=64)
parser.add_argument("--hidden_dim", type=int, default=16)
parser.add_argument("--hidden_dim_infer", type=int, default=16)
parser.add_argument("--cons_train", type=str, default="0.999_0.7")
parser.add_argument("--cons_test", type=str, default="0.999_0.001")
parser.add_argument("--num_classes", type=int, default=2)
parser.add_argument("--z_class_num", type=int, default=4)
parser.add_argument("--noise_ratio", type=float, default=0.1)
parser.add_argument("--penalty_anneal_iters", type=int, default=200)
parser.add_argument("--penalty_weight", type=float, default=10000.0)
parser.add_argument("--steps", type=int, default=501)
parser.add_argument("--scheduler", type=int, default=0)
flags = parser.parse_args()


def writeResult(flags, result: list):

    column_names = [
        "step",
        "iter",
        "lr",
        "Final train acc",
        "Final test acc",
        "Worst test acc",
        "Current step",
    ]
    filename = f"{flags.dataset}_{flags.irm_type}.xlsx"
    filePath = f"result/{filename}"

    try:

        wb = load_workbook(filename=filePath)
        ws = wb.active
    except FileNotFoundError:

        wb = Workbook()
        ws = wb.active

        for col, col_name in enumerate(column_names, 1):
            ws.cell(row=1, column=col).value = col_name

    start_row = ws.max_row + 1

    row = [
        flags.steps,
        flags.penalty_anneal_iters,
        flags.lr,
    ] + result

    for idx, value in enumerate(row):
        ws.cell(row=start_row, column=idx + 1, value=value)

    wb.save(filename=filePath)


## edited by w wang, rewrite parameter settings, for debug ONLY
# flagsl2_regularizer_weight = 0.001
# flags.lr = 0.1
# flags.noise_ratio = 0.2
# flags.cons_train = "0.999_0.9"
# flags.cons_test = "0.999_0.8_0.2_0.001"
# flags.penalty_weight = 10
# flags.steps = 400
# flags.dim_inv = 5
# flags.dim_sp = 5
# # flags.data_num_train = 39996
# # flags.data_num_test = 20000
# flags.n_restarts = 1
# flags.irm_type = "infer_irmv1_multi_class"
# flags.dataset = "landcover"
# flags.penalty_anneal_iters = 40
# flags.aux_num = 2
# flags.seed = 112
# flags.classes_num = 6
# flags.num_classes = 6
# flags.opt = "adam"
# flags.z_class_num = 2
# flags.scheduler = 1

print("batch_size is", flags.batch_size)

# torch.manual_seed(flags.seed)
if flags.dataset == "landcover":
    np.random.seed(flags.seed + 111)  # to be consistent with in-N-out.

flags.cons_ratio = "_".join([flags.cons_train, flags.cons_test])
flags.envs_num_train = len(flags.cons_train.split("_"))
flags.envs_num_test = len(flags.cons_test.split("_"))
assert flags.envs_num_test + flags.envs_num_train == len(flags.cons_ratio.split("_"))

for k, v in sorted(vars(flags).items()):
    print("\t{}: {}".format(k, v))

final_train_accs = []
final_test_accs = []

## add by w wang
final_test_accs_worst = []

for restart in range(flags.n_restarts):
    print("Restart", restart)

    (
        dp,
        mlp,
        test_batch_num,
        train_batch_num,
        val_batch_num,
        test_batch_fetcher,
        mean_nll,
        mean_accuracy,
        eval_acc,
    ) = init_dataset(flags)
    if mean_accuracy.__name__ == "mean_accuracy_reg":
        best_result_record = {
            "final": 10000,
            "worst": 10000,
            "final_epoch": -1,
            "worst_epoch": -1,
        }
    else:
        best_result_record = {
            "final": 0.0,
            "worst": 0,
            "final_epoch": -1,
            "worst_epoch": -1,
        }
    # if flags.opt == "adam":
    optimizer = optim.Adam(mlp.parameters(), lr=flags.lr)
    # elif flags.opt == "sgd":
    #     optimizer = optim.SGD(mlp.parameters(), momentum=0.9, lr=flags.lr)

    if flags.scheduler:
        scheduler = CosineLR(optimizer, flags.lr, flags.steps)

    scale = torch.ones(10).cuda().requires_grad_()
    algo = algorithm_builder(flags, dp)

    if flags.dataset == "house_price":
        meta_acc_test = MetaAcc(
            env=dp.envs_num_test, acc_measure=mean_accuracy, acc_type="test"
        )
    elif flags.dataset == "landcover":
        meta_acc_test = MetaAcc(env=2, acc_measure=mean_accuracy, acc_type="test")
    else:
        meta_acc_test = MetaAcc(env=0, acc_measure=mean_accuracy, acc_type="test")

    for step in tqdm(range(flags.steps)):
        mlp.train()
        for batch in range(train_batch_num):
            batch_data = dp.fetch_train()
            train_x, train_y, train_z, train_g, train_c, train_invnoise = batch_data

            # calculate train loss for different algorithms and datasets
            train_nll, train_penalty = algo(
                batch_data, step, mlp, scale, mean_nll=mean_nll
            )

            weight_norm = torch.tensor(0.0).cuda()
            for w in mlp.parameters():
                weight_norm += w.norm().pow(2)

            loss = train_nll.clone()
            loss += flags.l2_regularizer_weight * weight_norm
            penalty_weight = (
                flags.penalty_weight if step >= flags.penalty_anneal_iters else 0.0
            )
            if flags.irm_type == "erm":
                penalty_weight = 0
            train_penalty = torch.max(torch.tensor(0.0).cuda(), train_penalty.cuda())
            loss += penalty_weight * train_penalty
            if penalty_weight > 1.0:
                loss /= 1.0 + penalty_weight

            optimizer.zero_grad()
            loss.backward()
            optimizer.step()

        if flags.scheduler:
            scheduler.step()

        mlp.eval()
        train_accs = []
        meta_acc_test.clear()
        for ii in range(test_batch_num):
            test_x, test_y, test_z, test_g, test_c, test_invnoise = test_batch_fetcher()
            test_logits = mlp(test_x)
            meta_acc_test.process_batch(test_y, test_logits)
        meta_acc_test_res = meta_acc_test.meta_acc

        result = [
            0,
            np.mean(meta_acc_test_res["test_acc"].detach().cpu().numpy()),
            np.mean(meta_acc_test_res["test_acc_worst"].detach().cpu().numpy()),
            step,
        ]
        if mean_accuracy.__name__ == "mean_accuracy_reg":
            if result[1] < best_result_record["final"]:
                best_result_record["final"] = result[1]
                best_result_record["final_epoch"] = step
            if result[2] < best_result_record["worst"]:
                best_result_record["worst"] = result[2]
                best_result_record["worst_epoch"] = step
        else:
            if result[1] > best_result_record["final"]:
                best_result_record["final"] = result[1]
                best_result_record["final_epoch"] = step
            if result[2] > best_result_record["worst"]:
                best_result_record["worst"] = result[2]
                best_result_record["worst_epoch"] = step
        writeResult(flags, result)
    BEST_PATH = "best_result"
    bestResultPath = f"{BEST_PATH}/{flags.dataset}_{algorithmMap[flags.irm_type]}.csv"
    if not os.path.exists(bestResultPath):
        df = pd.DataFrame(
            {
                "final": [best_result_record["final"]],
                "worst": [best_result_record["worst"]],
                "final_epoch": [best_result_record["final_epoch"]],
                "worst_epoch": [best_result_record["worst_epoch"]],
            }
        )
        df.to_csv(bestResultPath, index=False)
    else:
        df = pd.read_csv(bestResultPath)
        new = pd.DataFrame(
            {
                "final": [best_result_record["final"]],
                "worst": [best_result_record["worst"]],
                "final_epoch": [best_result_record["final_epoch"]],
                "worst_epoch": [best_result_record["worst_epoch"]],
            }
        )
        df = pd.concat(
            [df, new],
            ignore_index=True,
        )
        df.to_csv(bestResultPath, index=False)

    # meta_test_acc_worst =
